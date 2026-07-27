import csv
import json
import os
import re
import tempfile
import unicodedata
import xml.etree.ElementTree as ET
import zipfile
from datetime import datetime, timezone
from urllib.parse import quote, urljoin
from urllib.request import Request, urlopen

from openpyxl import load_workbook

from manifest_utils import (
    build_file_entry,
    build_standard_manifest,
    utc_now_iso,
)


OUTPUT_ROOT = "countries"
USER_AGENT = "MostoFitNationalPackBuilder/1.0"
NL_NEVO_PAGE_URL = "https://www.rivm.nl/documenten/nevo-online-versie"
DE_BLS_SOURCE = "national_sources/DE/BLS_4_0_Daten_2025_DE.xlsx"

USDA_FDC_RELEASE_DATE = "2025-12-18"
USDA_FDC_VERSION = "14.0"
USDA_FDC_ZIP_URL = (
    "https://fdc.nal.usda.gov/fdc-datasets/"
    "FoodData_Central_foundation_food_json_2025-12-18.zip"
)
USDA_FDC_SOURCE_PAGE_URL = "https://fdc.nal.usda.gov/download-datasets/"
USDA_FDC_LICENSE_URL = "https://creativecommons.org/publicdomain/zero/1.0/"

CIQUAL_DATASET_DOI = "doi:10.57745/RDMHWY"
CIQUAL_API_BASE = "https://entrepot.recherche.data.gouv.fr/api"

CANADA_CNF_RESOURCE_ID = "019f2a90-e3a9-489d-b6e1-f74f4ba1d006"
OPEN_CANADA_API_BASE = "https://open.canada.ca/data/api/action"

GB_COFID_SOURCE = "national_sources/GB/McCance_Widdowsons_Composition_of_Foods_Integrated_Dataset_2021.xlsx"

AU_AFCD_FOOD_DETAILS_SOURCE = "national_sources/AU/AFCD Release 3 - Food Details.xlsx"
AU_AFCD_NUTRIENT_PROFILES_SOURCE = "national_sources/AU/AFCD Release 3 - Nutrient profiles.xlsx"
# ============================================================
# United States — USDA FoodData Central Foundation Foods
# ============================================================

def download_usda_fdc_json():
    temp_dir = tempfile.mkdtemp(prefix="usda_fdc_")
    zip_path = os.path.join(temp_dir, "usda_fdc.zip")

    print(f"Downloading USDA FoodData Central: {USDA_FDC_ZIP_URL}")
    download_file(USDA_FDC_ZIP_URL, zip_path)

    with zipfile.ZipFile(zip_path, "r") as archive:
        json_members = [
            member
            for member in archive.namelist()
            if member.lower().endswith(".json")
            and not member.endswith("/")
        ]

        if not json_members:
            raise RuntimeError(
                "The USDA FoodData Central archive contains no JSON file."
            )

        preferred_members = [
            member
            for member in json_members
            if "foundation_food" in os.path.basename(member).lower()
        ]

        candidates = preferred_members or json_members

        json_member = max(
            candidates,
            key=lambda member: archive.getinfo(member).file_size,
        )

        output_path = os.path.join(
            temp_dir,
            os.path.basename(json_member),
        )

        with archive.open(json_member) as source, open(output_path, "wb") as target:
            target.write(source.read())

    return output_path, os.path.basename(json_member)


def get_usda_nutrient_amount(food, nutrient_number):
    for food_nutrient in food.get("foodNutrients", []):
        nutrient = food_nutrient.get("nutrient") or {}

        number = normalize_text(nutrient.get("number"))

        if number != nutrient_number:
            continue

        return to_float(food_nutrient.get("amount"))

    return None


def build_usda_fdc():
    json_path, source_filename = download_usda_fdc_json()

    with open(json_path, "r", encoding="utf-8") as source_file:
        payload = json.load(source_file)

    source_foods = payload.get("FoundationFoods")

    if not isinstance(source_foods, list):
        raise RuntimeError(
            "USDA JSON does not contain a FoundationFoods array."
        )

    items = []
    seen_source_ids = set()

    for food in source_foods:
        source_item_id = normalize_text(food.get("fdcId"))
        name = normalize_text(food.get("description"))

        if not source_item_id or not name:
            continue

        if source_item_id in seen_source_ids:
            raise RuntimeError(
                f"Duplicate USDA fdcId found: {source_item_id}"
            )

        seen_source_ids.add(source_item_id)

        # USDA nutrient numbers:
        # 208 = Energy in kcal
        # 203 = Protein
        # 205 = Carbohydrate, by difference
        # 204 = Total lipid (fat)
        calories = get_usda_nutrient_amount(food, "208") or 0.0
        protein = get_usda_nutrient_amount(food, "203") or 0.0
        carbs = get_usda_nutrient_amount(food, "205") or 0.0
        fat = get_usda_nutrient_amount(food, "204") or 0.0

        if is_empty_macro_row(calories, protein, carbs, fat):
            continue

        item = make_pack_item(
            name=name,
            brand=None,
            calories=calories,
            protein=protein,
            carbs=carbs,
            fat=fat,
            source="usda_fdc",
            source_item_id=source_item_id,
        )

        items.append(item)

    items.sort(key=lambda item: item["name"].lower())

    national_path = os.path.join(
        OUTPUT_ROOT,
        "US",
        "national.json",
    )
    national_relative_path = "countries/US/national.json"
    manifest_path = os.path.join(
        OUTPUT_ROOT,
        "US",
        "national_manifest.json",
    )

    write_json(national_path, items)

    file_entry = build_file_entry(
        file_path=national_path,
        relative_path=national_relative_path,
        kind="national",
        record_count=len(items),
    )

    manifest = build_standard_manifest(
        pack_id="usda_fdc_us_foundation_14_0",
        pack_type="national",
        country_iso2="US",
        source="usda_fdc",
        source_name="FoodData Central Foundation Foods",
        publisher=(
            "U.S. Department of Agriculture, "
            "Agricultural Research Service"
        ),
        dataset_version=USDA_FDC_VERSION,
        license_id="CC0-1.0",
        source_url=USDA_FDC_ZIP_URL,
        license_url=USDA_FDC_LICENSE_URL,
        modified=True,
        modifications=[
            "selected Foundation Foods from the official USDA JSON release",
            "selected nutrient number 208 for energy in kilocalories",
            "selected nutrient number 203 for protein",
            "selected nutrient number 205 for carbohydrate by difference",
            "selected nutrient number 204 for total lipid",
            "set missing selected nutrient values to zero",
            "removed records with no selected energy or macronutrient values",
            "reduced records to the fields used by the MostoFit food schema",
            "preserved each USDA fdcId as sourceItemId",
            "normalized food names",
            "sorted records by food name",
            "converted the official USDA JSON release to app-facing JSON",
        ],
        generated_at=utc_now_iso(),
        record_count=len(items),
        files=[file_entry],
        extra_fields={
            "owner": (
                "U.S. Department of Agriculture, "
                "Agricultural Research Service"
            ),
            "releaseDate": USDA_FDC_RELEASE_DATE,
            "dataType": "Foundation Foods",
            "itemCount": len(items),
            "file": national_relative_path,
            "sourceFile": source_filename,
            "sourcePageUrl": USDA_FDC_SOURCE_PAGE_URL,
            "sourceArchiveUrl": USDA_FDC_ZIP_URL,
            "attribution": (
                "U.S. Department of Agriculture, "
                "Agricultural Research Service. FoodData Central."
            ),
        },
    )

    write_json(manifest_path, manifest)

    print(
        "Saved USDA FoodData Central Foundation Foods pack: "
        f"{len(items)} items"
    )

# ============================================================
# Shared helpers
# ============================================================

def normalize_text(value):
    return re.sub(r"\s+", " ", str(value or "").strip())


def normalize_for_match(value):
    text = normalize_text(value).lower()
    text = unicodedata.normalize("NFKD", text)
    return "".join(char for char in text if not unicodedata.combining(char))


def normalize_key(item):
    return normalize_text(item.get("name")).lower()


def to_float(value):
    if value in (None, ""):
        return None

    text = str(value).strip().replace(",", ".")
    text = re.sub(r"[^\d.\-]", "", text)

    if not text:
        return None

    try:
        return float(text)
    except ValueError:
        return None


def write_json(path, payload):
    os.makedirs(os.path.dirname(path), exist_ok=True)

    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def fetch_json(url):
    request = Request(url, headers={"User-Agent": USER_AGENT})
    with urlopen(request, timeout=60) as response:
        return json.loads(response.read().decode("utf-8"))


def download_file(url, output_path):
    request = Request(url, headers={"User-Agent": USER_AGENT})
    with urlopen(request, timeout=180) as response:
        with open(output_path, "wb") as f:
            f.write(response.read())
def fetch_text(url):
    request = Request(url, headers={"User-Agent": USER_AGENT})
    with urlopen(request, timeout=60) as response:
        return response.read().decode("utf-8", errors="replace")
    
def find_header(headers, starts_with):
    for header in headers:
        text = normalize_text(header)
        if text.startswith(starts_with):
            return header
    return None


def find_header_by_tokens(headers, *tokens):
    for header in headers:
        text = normalize_for_match(header)
        if all(normalize_for_match(token) in text for token in tokens):
            return header
    return None


def make_pack_item(
    name,
    brand,
    calories,
    protein,
    carbs,
    fat,
    source,
    source_item_id=None,
):
    item = {
        "name": normalize_text(name),
        "brand": brand,
        "barcode": None,
        "calories": round(calories, 2),
        "protein": round(protein, 2),
        "carbs": round(carbs, 2),
        "fat": round(fat, 2),
        "servingSize": 100.0,
        "servingUnit": "g",
        "isLiquid": False,
        "source": source,
    }

    normalized_source_item_id = normalize_text(source_item_id)

    if normalized_source_item_id:
        item["sourceItemId"] = normalized_source_item_id

    return item

def add_unique_item(items, seen, item):
    key = normalize_key(item)
    if key in seen:
        return

    seen.add(key)
    items.append(item)


def is_empty_macro_row(calories, protein, carbs, fat):
    return calories == 0.0 and protein == 0.0 and carbs == 0.0 and fat == 0.0

# ============================================================
# Netherlands — NEVO
# ============================================================

def find_nevo_zip_url(html):
    matches = re.findall(r'href=["\']([^"\']+\.zip(?:\?[^"\']*)?)["\']', html, flags=re.IGNORECASE)

    if not matches:
        raise RuntimeError("Could not find NEVO ZIP link on RIVM page.")

    return urljoin(NL_NEVO_PAGE_URL, matches[0])


def download_nevo_xlsx():
    html = fetch_text(NL_NEVO_PAGE_URL)
    zip_url = find_nevo_zip_url(html)

    temp_dir = tempfile.mkdtemp(prefix="nevo_")
    zip_path = os.path.join(temp_dir, "nevo.zip")

    print(f"Downloading Netherlands NEVO ZIP: {zip_url}")
    download_file(zip_url, zip_path)

    with zipfile.ZipFile(zip_path, "r") as zip_file:
        xlsx_members = [
            name for name in zip_file.namelist()
            if os.path.basename(name).lower().endswith(".xlsx")
        ]

        if not xlsx_members:
            raise RuntimeError("Could not find NEVO XLSX file in ZIP.")

        xlsx_members.sort()
        xlsx_member = xlsx_members[0]
        output_path = os.path.join(temp_dir, os.path.basename(xlsx_member))

        with zip_file.open(xlsx_member) as source, open(output_path, "wb") as target:
            target.write(source.read())

    return output_path, {
        "sourcePageUrl": NL_NEVO_PAGE_URL,
        "sourceZipUrl": zip_url,
        "sourceXlsxFile": os.path.basename(output_path),
    }


def build_netherlands_nevo():
    xlsx_path, source_info = download_nevo_xlsx()

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = workbook["NEVO2025"]

    rows = sheet.iter_rows(values_only=True)
    headers = list(next(rows))
    index = {header: i for i, header in enumerate(headers)}

    required_headers = {
        "NEVO-code",
        "Voedingsmiddelnaam/Dutch food name",
        "Engelse naam/Food name",
        "Hoeveelheid/Quantity",
        "ENERCC (kcal)",
        "PROT (g)",
        "CHO (g)",
        "FAT (g)",
    }

    missing = required_headers - set(headers)
    if missing:
        raise RuntimeError(f"Missing required NEVO headers: {missing}")

    items = []
    seen = set()

    for row in rows:
        code = row[index["NEVO-code"]]
        name_nl = normalize_text(row[index["Voedingsmiddelnaam/Dutch food name"]])
        name_en = normalize_text(row[index["Engelse naam/Food name"]])
        quantity = normalize_for_match(row[index["Hoeveelheid/Quantity"]])

        if not code or not name_nl:
            continue

        if quantity and quantity != "per 100g":
            continue

        calories = to_float(row[index["ENERCC (kcal)"]]) or 0.0
        protein = to_float(row[index["PROT (g)"]]) or 0.0
        carbs = to_float(row[index["CHO (g)"]]) or 0.0
        fat = to_float(row[index["FAT (g)"]]) or 0.0

        if is_empty_macro_row(calories, protein, carbs, fat):
            continue

        final_name = name_nl
        if name_en and name_en.lower() != name_nl.lower():
            final_name = f"{name_nl} / {name_en}"

        item = make_pack_item(
            name=final_name,
            brand="NEVO 2025",
            calories=calories,
            protein=protein,
            carbs=carbs,
            fat=fat,
            source="netherlands_nevo",
        )

        add_unique_item(items, seen, item)

    items.sort(key=lambda item: item["name"].lower())

    write_json(os.path.join(OUTPUT_ROOT, "NL", "national.json"), items)
    write_json(os.path.join(OUTPUT_ROOT, "NL", "national_manifest.json"), {
        "countryIso2": "NL",
        "source": "netherlands_nevo",
        "sourceName": "NEVO-online versie 2025 9.0",
        "owner": "RIVM",
        "license": "NEVO-online dataset conditions",
        "sourcePageUrl": source_info["sourcePageUrl"],
        "sourceZipUrl": source_info["sourceZipUrl"],
        "sourceXlsxFile": source_info["sourceXlsxFile"],
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "itemCount": len(items),
        "file": "countries/NL/national.json",
    })

    print(f"Saved Netherlands NEVO national pack: {len(items)} items")
# ============================================================
# Germany — BLS 4.0
# ============================================================

def build_germany_bls():
    if not os.path.exists(DE_BLS_SOURCE):
        raise FileNotFoundError(f"Missing BLS source file: {DE_BLS_SOURCE}")

    workbook = load_workbook(DE_BLS_SOURCE, read_only=True, data_only=True)
    sheet = workbook.active

    rows = sheet.iter_rows(values_only=True)
    headers = list(next(rows))
    header_set = set(headers)

    required_headers = {
        "BLS Code",
        "Lebensmittelbezeichnung",
        "Food name",
    }

    missing = required_headers - header_set
    if missing:
        raise RuntimeError(f"Missing required BLS headers: {missing}")

    kcal_header = find_header(headers, "ENERCC Energie (Kilokalorien)")
    protein_header = find_header(headers, "PROT625 Protein")
    fat_header = find_header(headers, "FAT Fett")
    carbs_header = find_header(headers, "CHO Kohlenhydrate, verfügbar")

    if not kcal_header or not protein_header or not fat_header or not carbs_header:
        raise RuntimeError(
            "Could not find required BLS nutrient columns. "
            f"kcal={kcal_header}, protein={protein_header}, fat={fat_header}, carbs={carbs_header}"
        )

    index = {header: i for i, header in enumerate(headers)}

    items = []
    seen = set()

    for row in rows:
        code = row[index["BLS Code"]]
        name_de = row[index["Lebensmittelbezeichnung"]]
        name_en = row[index["Food name"]]

        if not code or not name_de:
            continue

        calories = to_float(row[index[kcal_header]]) or 0.0
        protein = to_float(row[index[protein_header]]) or 0.0
        fat = to_float(row[index[fat_header]]) or 0.0
        carbs = to_float(row[index[carbs_header]]) or 0.0

        if is_empty_macro_row(calories, protein, carbs, fat):
            continue

        name = normalize_text(name_de)
        english_name = normalize_text(name_en)

        final_name = name
        if english_name and english_name.lower() != name.lower():
            final_name = f"{name} / {english_name}"

        item = make_pack_item(
            name=final_name,
            brand=None,
            calories=calories,
            protein=protein,
            carbs=carbs,
            fat=fat,
            source="germany_bls",
            source_item_id=code,
        )

        add_unique_item(items, seen, item)

    items.sort(key=lambda item: item["name"].lower())

    national_path = os.path.join(OUTPUT_ROOT, "DE", "national.json")
    national_relative_path = "countries/DE/national.json"
    manifest_path = os.path.join(
        OUTPUT_ROOT,
        "DE",
        "national_manifest.json",
    )

    write_json(national_path, items)

    file_entry = build_file_entry(
        file_path=national_path,
        relative_path=national_relative_path,
        kind="national",
        record_count=len(items),
    )

    manifest = build_standard_manifest(
        pack_id="bls_de_4_0",
        pack_type="national",
        country_iso2="DE",
        source="germany_bls",
        source_name=(
            "Bundeslebensmittelschlüssel (BLS), Version 4.0 "
            "- Deutsche Nährstoffdatenbank"
        ),
        publisher="Max Rubner-Institut",
        dataset_version="4.0",
        license_id="CC-BY-4.0",
        source_url="https://blsdb.de/download",
        license_url="https://creativecommons.org/licenses/by/4.0/",
        modified=True,
        modifications=[
            "selected energy, protein, carbohydrate and fat fields",
            "combined German and English food names where available",
            "normalized text and nutrient values to the MostoFit food schema",
            "removed records with no energy or macronutrient values",
            "deduplicated records by normalized food name",
            "sorted records by food name",
            "converted the source XLSX workbook to JSON",
        ],
        generated_at=utc_now_iso(),
        record_count=len(items),
        files=[file_entry],
        extra_fields={
            "owner": "Max Rubner-Institut",
            "itemCount": len(items),
            "file": national_relative_path,
            "sourceFile": DE_BLS_SOURCE,
            "doi": "10.25826/Data20251217-134202-0",
            "attribution": (
                "Max Rubner-Institut (2025): "
                "Bundeslebensmittelschlüssel (BLS), Version 4.0 "
                "- Deutsche Nährstoffdatenbank. Karlsruhe."
            ),
        },
    )

    write_json(manifest_path, manifest)
    print(f"Saved Germany BLS national pack: {len(items)} items")


# ============================================================
# France — CIQUAL
# ============================================================

def get_ciqual_dataset_files():
    encoded_doi = quote(CIQUAL_DATASET_DOI, safe=":")
    url = (
        f"{CIQUAL_API_BASE}/datasets/export"
        f"?exporter=dataverse_json&persistentId={encoded_doi}"
    )

    payload = fetch_json(url)
    return payload.get("datasetVersion", {}).get("files", [])


def find_ciqual_file(files, starts_with, ends_with=".xml"):
    matches = []

    for file_entry in files:
        data_file = file_entry.get("dataFile", {})
        filename = data_file.get("filename", "")
        persistent_id = data_file.get("persistentId", "")
        filename_l = filename.lower()

        if not persistent_id:
            continue

        if not filename_l.endswith(ends_with):
            continue

        if starts_with == "alim_":
            if filename_l.startswith("alim_") and not filename_l.startswith("alim_grp_"):
                matches.append({
                    "filename": filename,
                    "persistentId": persistent_id,
                })
            continue

        if filename_l.startswith(starts_with):
            matches.append({
                "filename": filename,
                "persistentId": persistent_id,
            })

    if not matches:
        available = [
            file_entry.get("dataFile", {}).get("filename", "")
            for file_entry in files
        ]
        raise RuntimeError(
            f"Could not find CIQUAL file starting with '{starts_with}'. "
            f"Available files: {available}"
        )

    matches.sort(key=lambda item: item["filename"])

    if starts_with == "alim_":
        for item in matches:
            basename = os.path.basename(item["filename"]).lower()
            if basename.startswith("alim_") and not basename.startswith("alim_grp_"):
                return item

    return matches[-1]


def download_ciqual_xml_files():
    files = get_ciqual_dataset_files()
    temp_dir = tempfile.mkdtemp(prefix="ciqual_")

    wanted = {
        "alim": find_ciqual_file(files, "alim_"),
        "const": find_ciqual_file(files, "const_"),
        "compo": find_ciqual_file(files, "compo_"),
    }

    downloaded = {}

    for key, file_info in wanted.items():
        encoded_file_doi = quote(file_info["persistentId"], safe=":")
        download_url = (
            f"{CIQUAL_API_BASE}/access/datafile/:persistentId"
            f"?persistentId={encoded_file_doi}"
        )

        output_path = os.path.join(temp_dir, file_info["filename"])

        print(f"Downloading CIQUAL {key} file: {file_info['filename']}")
        download_file(download_url, output_path)

        downloaded[key] = {
            "path": output_path,
            "filename": file_info["filename"],
            "persistentId": file_info["persistentId"],
        }

    return downloaded


def xml_text(parent, tag_name):
    child = parent.find(tag_name)
    if child is None or child.text is None:
        return ""
    return normalize_text(child.text)


def parse_ciqual_alim(path):
    foods = {}

    root = ET.parse(path).getroot()

    for alim in root.findall("ALIM"):
        code = xml_text(alim, "alim_code")
        name_fr = xml_text(alim, "alim_nom_fr")
        name_en = xml_text(alim, "alim_nom_eng")

        if not code or not name_fr:
            continue

        final_name = name_fr
        if name_en and name_en.lower() != name_fr.lower():
            final_name = f"{name_fr} / {name_en}"

        foods[code] = final_name

    return foods


def parse_ciqual_const(path):
    nutrients = {}

    root = ET.parse(path).getroot()

    for const in root.findall("CONST"):
        code = xml_text(const, "const_code")
        name_fr = xml_text(const, "const_nom_fr")
        name_en = xml_text(const, "const_nom_eng")
        infoods = xml_text(const, "code_INFOODS")

        if code:
            nutrients[code] = {
                "name_fr": name_fr,
                "name_en": name_en,
                "infoods": infoods,
            }

    return nutrients


def identify_ciqual_macro_codes(nutrients):
    kcal_code = None
    protein_code = None
    fat_code = None
    carbs_code = None

    for code, meta in nutrients.items():
        infoods = normalize_for_match(meta.get("infoods", ""))
        name = normalize_for_match(
            " ".join([
                meta.get("name_fr", ""),
                meta.get("name_en", ""),
                meta.get("infoods", ""),
            ])
        )

        if not kcal_code and "kcal" in name:
            kcal_code = code

        if not protein_code and (
            "proteines" in name or "protein" in name or infoods == "prot"
        ):
            protein_code = code

        if not fat_code and (
            "lipides" in name or "fat" in name or infoods == "fat"
        ):
            fat_code = code

        if not carbs_code and (
            "glucides" in name
            or "carbohydrate" in name
            or infoods in {"choavl", "chocdf"}
        ):
            carbs_code = code

    if not kcal_code or not protein_code or not fat_code or not carbs_code:
        raise RuntimeError(
            "Could not identify CIQUAL macro codes. "
            f"kcal={kcal_code}, protein={protein_code}, fat={fat_code}, carbs={carbs_code}"
        )

    print(
        "CIQUAL macro codes: "
        f"kcal={kcal_code}, protein={protein_code}, fat={fat_code}, carbs={carbs_code}"
    )

    return {
        "calories": kcal_code,
        "protein": protein_code,
        "fat": fat_code,
        "carbs": carbs_code,
    }


def parse_ciqual_compo(path, macro_codes):
    wanted_codes = set(macro_codes.values())
    reverse_codes = {value: key for key, value in macro_codes.items()}

    values_by_food = {}

    root = ET.parse(path).getroot()

    parsed_records = 0
    parsed_macro_values = 0

    for compo in root.findall("COMPO"):
        food_code = xml_text(compo, "alim_code")
        nutrient_code = xml_text(compo, "const_code")
        raw_value = xml_text(compo, "teneur")

        if not food_code or not nutrient_code:
            continue

        parsed_records += 1

        if nutrient_code not in wanted_codes:
            continue

        value = to_float(raw_value)
        if value is None:
            continue

        field = reverse_codes[nutrient_code]
        values_by_food.setdefault(food_code, {})[field] = value
        parsed_macro_values += 1

    print(
        "CIQUAL composition parsed: "
        f"records={parsed_records}, macroValues={parsed_macro_values}, foodsWithMacros={len(values_by_food)}"
    )

    return values_by_food


def build_france_ciqual():
    ciqual_files = download_ciqual_xml_files()

    foods = parse_ciqual_alim(ciqual_files["alim"]["path"])
    nutrients = parse_ciqual_const(ciqual_files["const"]["path"])
    macro_codes = identify_ciqual_macro_codes(nutrients)
    values_by_food = parse_ciqual_compo(ciqual_files["compo"]["path"], macro_codes)

    items = []
    seen = set()

    for food_code, name in foods.items():
        values = values_by_food.get(food_code, {})

        calories = values.get("calories", 0.0)
        protein = values.get("protein", 0.0)
        fat = values.get("fat", 0.0)
        carbs = values.get("carbs", 0.0)

        if is_empty_macro_row(calories, protein, carbs, fat):
            continue

        item = make_pack_item(
            name=name,
            brand="CIQUAL",
            calories=calories,
            protein=protein,
            carbs=carbs,
            fat=fat,
            source="france_ciqual",
        )

        add_unique_item(items, seen, item)

    items.sort(key=lambda item: item["name"].lower())

    write_json(os.path.join(OUTPUT_ROOT, "FR", "national.json"), items)
    write_json(os.path.join(OUTPUT_ROOT, "FR", "national_manifest.json"), {
        "countryIso2": "FR",
        "source": "france_ciqual",
        "sourceName": "Table de composition nutritionnelle des aliments Ciqual",
        "owner": "ANSES",
        "license": "Etalab Open License 2.0",
        "datasetPersistentId": CIQUAL_DATASET_DOI,
        "sourceFiles": {
            "alim": {
                "filename": ciqual_files["alim"]["filename"],
                "persistentId": ciqual_files["alim"]["persistentId"],
            },
            "const": {
                "filename": ciqual_files["const"]["filename"],
                "persistentId": ciqual_files["const"]["persistentId"],
            },
            "compo": {
                "filename": ciqual_files["compo"]["filename"],
                "persistentId": ciqual_files["compo"]["persistentId"],
            },
        },
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "itemCount": len(items),
        "file": "countries/FR/national.json",
    })

    print(f"Saved France CIQUAL national pack: {len(items)} items")


# ============================================================
# Canada — CNF 2026
# ============================================================

def download_canada_cnf_zip():
    url = f"{OPEN_CANADA_API_BASE}/resource_show?id={CANADA_CNF_RESOURCE_ID}"
    payload = fetch_json(url)

    result = payload.get("result", {})
    download_url = result.get("url")
    filename = result.get("name") or result.get("title") or "canada_cnf_2026.zip"

    if not download_url:
        raise RuntimeError("Could not find Canada CNF download URL from Open Canada API.")

    temp_dir = tempfile.mkdtemp(prefix="canada_cnf_")
    output_path = os.path.join(temp_dir, "canada_cnf_2026.zip")

    print(f"Downloading Canada CNF ZIP: {filename}")
    download_file(download_url, output_path)

    return output_path, {
        "resourceId": CANADA_CNF_RESOURCE_ID,
        "sourceUrl": download_url,
        "filename": filename,
    }


def find_zip_csv(zip_file, filename):
    target = filename.lower()

    for name in zip_file.namelist():
        if os.path.basename(name).lower() == target:
            return name

    raise RuntimeError(f"Could not find {filename} in ZIP")


def read_csv_from_zip(zip_file, member):
    with zip_file.open(member) as f:
        text = f.read().decode("utf-8-sig", errors="replace").splitlines()
    return list(csv.DictReader(text))


def build_canada_cnf():
    zip_path, source_info = download_canada_cnf_zip()

    with zipfile.ZipFile(zip_path, "r") as zip_file:
        food_rows = read_csv_from_zip(zip_file, find_zip_csv(zip_file, "Food_Name.csv"))
        nutrient_rows = read_csv_from_zip(zip_file, find_zip_csv(zip_file, "Nutrient_Amount.csv"))

    foods = {}

    for row in food_rows:
        code = normalize_text(row.get("Food_Code"))
        name_en = normalize_text(row.get("Food_Description_EN"))
        name_fr = normalize_text(row.get("Food_Description_FR"))

        if not code or not name_en:
            continue

        final_name = name_en
        if name_fr and name_fr.lower() != name_en.lower():
            final_name = f"{name_en} / {name_fr}"

        foods[code] = final_name

    macro_codes = {
        "protein": "203",
        "fat": "204",
        "carbs": "205",
        "calories": "208",
    }

    values_by_food = {}

    for row in nutrient_rows:
        food_code = normalize_text(row.get("Food_Code"))
        nutrient_code = normalize_text(row.get("Nutrient_Code"))

        if not food_code or nutrient_code not in macro_codes.values():
            continue

        value = to_float(row.get("Nutrient_Amount"))
        if value is None:
            continue

        field = next(
            field_name
            for field_name, code in macro_codes.items()
            if code == nutrient_code
        )

        values_by_food.setdefault(food_code, {})[field] = value

    items = []
    seen = set()

    for food_code, name in foods.items():
        values = values_by_food.get(food_code, {})

        calories = values.get("calories", 0.0)
        protein = values.get("protein", 0.0)
        fat = values.get("fat", 0.0)
        carbs = values.get("carbs", 0.0)

        if is_empty_macro_row(calories, protein, carbs, fat):
            continue

        item = make_pack_item(
            name=name,
            brand="CNF 2026",
            calories=calories,
            protein=protein,
            carbs=carbs,
            fat=fat,
            source="canada_cnf",
        )

        add_unique_item(items, seen, item)

    items.sort(key=lambda item: item["name"].lower())

    write_json(os.path.join(OUTPUT_ROOT, "CA", "national.json"), items)
    write_json(os.path.join(OUTPUT_ROOT, "CA", "national_manifest.json"), {
        "countryIso2": "CA",
        "source": "canada_cnf",
        "sourceName": "Canadian Nutrient File 2026",
        "owner": "Health Canada",
        "license": "Open Government Licence - Canada",
        "resourceId": source_info["resourceId"],
        "sourceUrl": source_info["sourceUrl"],
        "sourceFileName": source_info["filename"],
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "itemCount": len(items),
        "file": "countries/CA/national.json",
    })

    print(f"Saved Canada CNF national pack: {len(items)} items")


# ============================================================
# United Kingdom — CoFID 2021
# ============================================================

def build_united_kingdom_cofid():
    if not os.path.exists(GB_COFID_SOURCE):
        raise FileNotFoundError(f"Missing CoFID source file: {GB_COFID_SOURCE}")

    workbook = load_workbook(GB_COFID_SOURCE, read_only=True, data_only=True)
    sheet = workbook["1.3 Proximates"]

    rows = sheet.iter_rows(values_only=True)
    headers = list(next(rows))
    index = {header: i for i, header in enumerate(headers)}

    required_headers = {
        "Food Code",
        "Food Name",
        "Protein (g)",
        "Fat (g)",
        "Carbohydrate (g)",
        "Energy (kcal) (kcal)",
    }

    missing = required_headers - set(headers)
    if missing:
        raise RuntimeError(f"Missing required CoFID headers: {missing}")

    items = []
    seen = set()

    next(rows, None)
    next(rows, None)

    for row in rows:
        code = row[index["Food Code"]]
        name = row[index["Food Name"]]

        if not code or not name:
            continue

        calories = to_float(row[index["Energy (kcal) (kcal)"]]) or 0.0
        protein = to_float(row[index["Protein (g)"]]) or 0.0
        fat = to_float(row[index["Fat (g)"]]) or 0.0
        carbs = to_float(row[index["Carbohydrate (g)"]]) or 0.0

        if is_empty_macro_row(calories, protein, carbs, fat):
            continue

        item = make_pack_item(
            name=name,
            brand="CoFID 2021",
            calories=calories,
            protein=protein,
            carbs=carbs,
            fat=fat,
            source="uk_cofid",
        )

        add_unique_item(items, seen, item)

    items.sort(key=lambda item: item["name"].lower())

    write_json(os.path.join(OUTPUT_ROOT, "GB", "national.json"), items)
    write_json(os.path.join(OUTPUT_ROOT, "GB", "national_manifest.json"), {
        "countryIso2": "GB",
        "source": "uk_cofid",
        "sourceName": "McCance and Widdowson's Composition of Foods Integrated Dataset 2021",
        "owner": "UK Government / Public Health England",
        "license": "Open Government Licence",
        "sourceFile": GB_COFID_SOURCE,
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "itemCount": len(items),
        "file": "countries/GB/national.json",
    })

    print(f"Saved UK CoFID national pack: {len(items)} items")


# ============================================================
# Australia — AFCD Release 3
# ============================================================

def build_australia_afcd():
    if not os.path.exists(AU_AFCD_FOOD_DETAILS_SOURCE):
        raise FileNotFoundError(f"Missing AFCD food details file: {AU_AFCD_FOOD_DETAILS_SOURCE}")

    if not os.path.exists(AU_AFCD_NUTRIENT_PROFILES_SOURCE):
        raise FileNotFoundError(f"Missing AFCD nutrient profiles file: {AU_AFCD_NUTRIENT_PROFILES_SOURCE}")

    food_workbook = load_workbook(AU_AFCD_FOOD_DETAILS_SOURCE, read_only=True, data_only=True)
    food_sheet = food_workbook["Food details"]

    food_rows = food_sheet.iter_rows(values_only=True)
    next(food_rows, None)
    next(food_rows, None)

    food_headers = list(next(food_rows))
    food_index = {header: i for i, header in enumerate(food_headers)}

    food_key_header = find_header_by_tokens(food_headers, "public", "food", "key")
    food_name_header = find_header_by_tokens(food_headers, "food", "name")

    if not food_key_header or not food_name_header:
        raise RuntimeError(
            "Could not find required AFCD food columns. "
            f"foodKey={food_key_header}, foodName={food_name_header}"
        )

    food_names = {}

    for row in food_rows:
        food_key = normalize_text(row[food_index[food_key_header]])
        food_name = normalize_text(row[food_index[food_name_header]])

        if food_key and food_name:
            food_names[food_key] = food_name

    nutrient_workbook = load_workbook(AU_AFCD_NUTRIENT_PROFILES_SOURCE, read_only=True, data_only=True)
    nutrient_sheet = nutrient_workbook["All solids & liquids per 100 g"]

    nutrient_rows = nutrient_sheet.iter_rows(values_only=True)
    next(nutrient_rows, None)
    next(nutrient_rows, None)

    nutrient_headers = list(next(nutrient_rows))
    nutrient_index = {header: i for i, header in enumerate(nutrient_headers)}

    nutrient_food_key_header = find_header_by_tokens(nutrient_headers, "public", "food", "key")
    nutrient_food_name_header = find_header_by_tokens(nutrient_headers, "food", "name")
    kj_header = find_header_by_tokens(nutrient_headers, "energy", "without dietary fibre", "kj")
    protein_header = find_header_by_tokens(nutrient_headers, "protein")
    fat_header = find_header_by_tokens(nutrient_headers, "fat", "total")
    carbs_header = find_header_by_tokens(nutrient_headers, "carbohydrate")

    if (
        not nutrient_food_key_header
        or not nutrient_food_name_header
        or not kj_header
        or not protein_header
        or not fat_header
        or not carbs_header
    ):
        raise RuntimeError(
            "Could not find required AFCD nutrient columns. "
            f"foodKey={nutrient_food_key_header}, foodName={nutrient_food_name_header}, "
            f"kj={kj_header}, protein={protein_header}, fat={fat_header}, carbs={carbs_header}"
        )

    items = []
    seen = set()

    for row in nutrient_rows:
        food_key = normalize_text(row[nutrient_index[nutrient_food_key_header]])

        if not food_key:
            continue

        name = food_names.get(food_key) or normalize_text(row[nutrient_index[nutrient_food_name_header]])

        if not name:
            continue

        energy_kj = to_float(row[nutrient_index[kj_header]]) or 0.0
        calories = energy_kj / 4.184
        protein = to_float(row[nutrient_index[protein_header]]) or 0.0
        fat = to_float(row[nutrient_index[fat_header]]) or 0.0
        carbs = to_float(row[nutrient_index[carbs_header]]) or 0.0

        if is_empty_macro_row(calories, protein, carbs, fat):
            continue

        item = make_pack_item(
            name=name,
            brand="AFCD Release 3",
            calories=calories,
            protein=protein,
            carbs=carbs,
            fat=fat,
            source="australia_afcd",
        )

        add_unique_item(items, seen, item)

    items.sort(key=lambda item: item["name"].lower())

    write_json(os.path.join(OUTPUT_ROOT, "AU", "national.json"), items)
    write_json(os.path.join(OUTPUT_ROOT, "AU", "national_manifest.json"), {
        "countryIso2": "AU",
        "source": "australia_afcd",
        "sourceName": "Australian Food Composition Database Release 3",
        "owner": "Food Standards Australia New Zealand",
        "license": "Food Standards Australia New Zealand data files",
        "sourceFiles": {
            "foodDetails": AU_AFCD_FOOD_DETAILS_SOURCE,
            "nutrientProfiles": AU_AFCD_NUTRIENT_PROFILES_SOURCE,
        },
        "energyConversion": "kcal = kJ / 4.184",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "itemCount": len(items),
        "file": "countries/AU/national.json",
    })

    print(f"Saved Australia AFCD national pack: {len(items)} items")
# ============================================================
# Entry point
# ============================================================

def main():
    build_germany_bls()
    build_france_ciqual()
    build_usda_fdc()
    build_canada_cnf()
    build_united_kingdom_cofid()
    build_australia_afcd()
    build_netherlands_nevo()


if __name__ == "__main__":
    main()
